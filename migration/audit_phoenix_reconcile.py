#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Por cada **fila de artículo** (cada código en la rejilla), compara el stock **simulado**
(INITIAL + deltas en las **columnas de movimiento** + snapshots INITIAL QUANTITY intermedios)
con el **inventario** en la columna **INITIAL QUANTITY** de inventario (**más a la izquierda** / más reciente).

Salida: phoenix_reconcile_mismatches.csv, phoenix_reconcile_mal_leidas.csv, phoenix_reconcile_ops_descuadre.csv,
phoenix_reconcile_diagnostico_filas.txt (errores explicados por fila), phoenix_reconcile_by_column.txt

Uso:
  py -3 migration/audit_phoenix_reconcile.py -i Libro1.xlsx
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl: py -3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from phoenix_sheet_utils import (
    diagnostico_texto_filas_problematicas,
    find_last_initial_quantity_column,
    run_reconciliation_audit,
)


def find_header_row(ws, max_scan: int = 30) -> int:
    for r in range(1, max_scan + 1):
        v = ws.cell(r, 3).value
        if v is None:
            continue
        if str(v).strip().upper() == "CODE":
            return r
    raise ValueError("No se encontró fila con 'CODE' en columna C.")


def column_dates(ws, date_row: int) -> dict[int, str | None]:
    from datetime import date, datetime

    cur: str | None = None
    out: dict[int, str | None] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(date_row, c).value
        if isinstance(v, datetime):
            cur = v.strftime("%Y-%m-%d")
        elif isinstance(v, date):
            cur = v.isoformat()
        out[c] = cur
    return out


def discover_date_row(ws, header_row: int) -> int:
    if header_row > 1:
        return header_row - 1
    return 8


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Auditoría por filas de artículo: movimientos (columnas) vs inventario"
    )
    ap.add_argument("-i", "--input", required=True)
    ap.add_argument("--sheet", default="HISTORY")
    ap.add_argument("--header-row", type=int, default=0)
    ap.add_argument("--date-row", type=int, default=0)
    ap.add_argument("--code-col", type=int, default=3)
    ap.add_argument("--initial-col", type=int, default=7)
    ap.add_argument(
        "--last-initial-col",
        type=int,
        default=0,
        help="Columna Excel de INITIAL QUANTITY para inventario (0=auto: la más a la izquierda/más reciente)",
    )
    ap.add_argument("--first-move-col", type=int, default=9)
    ap.add_argument("--audit-tolerance", type=float, default=0.5)
    ap.add_argument(
        "--out-dir",
        default="",
        help="Default: migration/generated",
    )
    args = ap.parse_args()

    migration_dir = Path(__file__).resolve().parent
    out_dir = Path(args.out_dir).resolve() if args.out_dir else (migration_dir / "generated")
    out_dir.mkdir(parents=True, exist_ok=True)

    in_path = Path(args.input).expanduser().resolve()
    if not in_path.is_file():
        print(f"No existe: {in_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(in_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        print("Hojas:", wb.sheetnames, file=sys.stderr)
        sys.exit(1)
    ws = wb[args.sheet]

    header_row = args.header_row if args.header_row > 0 else find_header_row(ws)
    date_row = args.date_row if args.date_row > 0 else discover_date_row(ws, header_row)
    first_data = header_row + 1
    if args.last_initial_col > 0:
        last_iq = args.last_initial_col
    else:
        last_iq = find_last_initial_quantity_column(ws, header_row)
        if last_iq is None:
            print("No hay columna INITIAL_QUANTITY; usa --last-initial-col", file=sys.stderr)
            sys.exit(1)

    col_dt = column_dates(ws, date_row)
    (
        mism,
        col_hits,
        total_art,
        ok_n,
        mal_leidas,
        ops_descuadre_todas,
    ) = run_reconciliation_audit(
        ws,
        header_row,
        first_data,
        last_iq,
        args.first_move_col,
        args.initial_col,
        args.code_col,
        col_dt,
        abs_tol=args.audit_tolerance,
    )

    audit_csv = out_dir / "phoenix_reconcile_mismatches.csv"
    with audit_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "fila",
                "codigo",
                "initial_col",
                "calculado",
                "inventario",
                "diff",
                "movimientos_col_delta",
                "detalle_modos_parseo",
                "alertas_lectura",
            ],
        )
        w.writeheader()
        w.writerows(mism)

    audit_mal = out_dir / "phoenix_reconcile_mal_leidas.csv"
    with audit_mal.open("w", encoding="utf-8-sig", newline="") as f:
        wm = csv.DictWriter(
            f,
            fieldnames=[
                "fila",
                "codigo",
                "col_movimiento",
                "delta",
                "modo_parseo",
                "trozo_celda",
                "motivo",
            ],
        )
        wm.writeheader()
        wm.writerows(mal_leidas)

    audit_ops = out_dir / "phoenix_reconcile_ops_descuadre.csv"
    with audit_ops.open("w", encoding="utf-8-sig", newline="") as f:
        wo = csv.DictWriter(
            f,
            fieldnames=[
                "fila",
                "codigo",
                "col_movimiento",
                "delta",
                "modo_parseo",
                "trozo_celda",
                "diff_inventario_fila",
            ],
        )
        wo.writeheader()
        wo.writerows(ops_descuadre_todas)

    diag_path = out_dir / "phoenix_reconcile_diagnostico_filas.txt"
    diag_path.write_text(
        diagnostico_texto_filas_problematicas(mism, mal_leidas, ops_descuadre_todas),
        encoding="utf-8",
    )

    audit_txt = out_dir / "phoenix_reconcile_by_column.txt"
    lines = [
        "# Columnas de MOVIMIENTO donde hubo ____ en FILAS DE ARTÍCULO descuadradas.",
        f"# Filas de artículo: {total_art} | OK: {ok_n} | descuadradas: {len(mism)}",
        f"# INITIAL QUANTITY de inventario (más a la izquierda / más reciente): columna {last_iq}",
        "",
    ]
    for c, n in col_hits.most_common():
        lines.append(f"col {c}\t{n}")
    audit_txt.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(audit_csv)
    print(audit_mal)
    print(audit_ops)
    print(diag_path)
    print(audit_txt)
    print(f"Resumen: {ok_n}/{total_art} filas de artículo coinciden con inventario.")
    print(f"Operaciones con lectura sospechosa: {len(mal_leidas)}")


if __name__ == "__main__":
    main()
