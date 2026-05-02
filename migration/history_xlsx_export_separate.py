#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera **dos archivos separados** desde la hoja Phoenix HISTORY:

1. **Inventario** — CSV listo para Configuración → Importar CSV (solo inventario).
2. **Respaldo** — JSON completo G-NEEX (inventario + movimientos + claves por defecto)
   listo para Configuración → Importar respaldo.

El JSON incluye inventario embebido (mismo **stock** que el CSV) para que los
`itemId` de los movimientos coincidan. El **StockPrincipal** sale de la **primera**
subcabecera **INITIAL QUANTITY** al ir **de izquierda a derecha** (en esta hoja lo
**más reciente** está a la izquierda; no del último movimiento de la fila).
El CSV es la copia “solo inventario” para flujos donde quieras ese archivo aparte.

**Modelo Phoenix:** cada **columna** (desde la primera columna de movimientos) es **un
solo movimiento**; todas las **filas** hacia abajo son **líneas de artículo** dentro
de ese movimiento (celdas con texto `____`). El **tipo** prioriza el **color de relleno**
de esas celdas (`cellFillRgbToType` en `phoenix-type-map.json`); si no hay color
mapeado, se usa la cabecera de columna.

Tras generar intermediate.json, invoca Node: `build-gneex-backup.mjs`.

Uso:
  py -3 migration/history_xlsx_export_separate.py --input "%USERPROFILE%\\Desktop\\Libro1.xlsx"

Opcional:
  py -3 migration/history_xlsx_export_separate.py -i libro.xlsx --no-node
    (solo CSV + intermediate.json; ejecuta Node tú mismo)

Requiere: pip install openpyxl | Node.js para el JSON final
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
from collections import Counter
from colorsys import hls_to_rgb, rgb_to_hls, rgb_to_hsv
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    import openpyxl.styles.colors as XL_COLORS
except ImportError:
    print("Instala openpyxl: py -3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from gneex_inventory_csv_merge import (
    merge_inv_list_from_gneex_csv,
    merge_inv_list_stocks_from_gneex_backup,
)
from phoenix_sheet_utils import (
    anchor_cell,
    diagnostico_texto_filas_problematicas,
    effective_cell_value,
    find_last_initial_quantity_column,
    parse_movement_cell,
    run_reconciliation_audit,
    stock_from_initial_quantity_cell,
)

# Mismo orden que js/inventory.js INVENTORY_IMPORT_CSV_HEADERS
CSV_HEADERS = [
    "Codigo",
    "Descripcion",
    "Categoria",
    "StockPrincipal",
    "StockProduccion",
    "StockTransformacion",
    "CantidadPorCaja",
    "NumeroCajas",
    "Ubicacion",
    "FechaExpedicion",
    "DiasParaExpirar",
    "FechaExpiracion",
    "Proveedor",
    "UltimaOrden",
    "Detalles",
    "Id",
    "StockMinimo",
    "StockMaximo",
    "VidaUtilMeses",
    "Notas",
    "LotesJson",
]

def find_header_row(ws, max_scan: int = 30) -> int:
    for r in range(1, max_scan + 1):
        v = ws.cell(r, 3).value
        if v is None:
            continue
        if str(v).strip().upper() == "CODE":
            return r
    raise ValueError("No se encontró fila con 'CODE' en columna C.")


def column_dates(ws, date_row: int) -> dict[int, str | None]:
    """Fecha YYYY-MM-DD por columna (propagada hacia la derecha desde celdas fecha)."""
    cur: str | None = None
    out: dict[int, str | None] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(date_row, c).value
        if isinstance(v, datetime):
            cur = v.strftime("%Y-%m-%d")
        elif isinstance(v, date):
            cur = v.isoformat()
        out[c] = cur
    return out


def load_type_map(migration_dir: Path) -> tuple[dict[str, str], str, dict[str, str]]:
    p = migration_dir / "phoenix-type-map.json"
    if not p.is_file():
        return {}, "AJUSTE", {}
    j = json.loads(p.read_text(encoding="utf-8"))
    m = {}
    for k, v in (j.get("map") or {}).items():
        if isinstance(k, str) and isinstance(v, str):
            m[k.strip().upper().replace(" ", "_")] = v
    fill_map: dict[str, str] = {}
    for k, v in (j.get("cellFillRgbToType") or {}).items():
        if not isinstance(k, str) or not isinstance(v, str):
            continue
        kk = k.strip().upper().replace("#", "")
        if len(kk) == 8 and all(c in "0123456789ABCDEF" for c in kk):
            kk = kk[2:8]
        if len(kk) == 6 and all(c in "0123456789ABCDEF" for c in kk):
            fill_map[kk] = v
    return m, str(j.get("defaultType") or "AJUSTE"), fill_map


# --- Color de relleno: RGB directo, indexado o tema (+ tinte Excel) ---
_RGBMAX = 255
_HLSMAX = 240.0


def _rgb_hex_to_ms_hls(rgb_hex: str) -> tuple[int, int, int]:
    h = rgb_hex[-6:]
    r = int(h[0:2], 16) / _RGBMAX
    g = int(h[2:4], 16) / _RGBMAX
    b = int(h[4:6], 16) / _RGBMAX
    hh, ll, ss = rgb_to_hls(r, g, b)
    return (
        int(round(hh * _HLSMAX)),
        int(round(ll * _HLSMAX)),
        int(round(ss * _HLSMAX)),
    )


def _ms_hls_to_rgb_hex(h: int, l: int, s: int) -> str:
    t = hls_to_rgb(h / _HLSMAX, l / _HLSMAX, s / _HLSMAX)
    return (
        "%02X%02X%02X"
        % (
            int(round(t[0] * _RGBMAX)),
            int(round(t[1] * _RGBMAX)),
            int(round(t[2] * _RGBMAX)),
        )
    )


def _tint_luminance(tint: float, lum: int) -> int:
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    return int(round(lum * (1.0 - tint) + (_HLSMAX - _HLSMAX * (1.0 - tint))))


def _theme_base_and_tint_to_rrggbb(base_rrggbb: str, tint: float) -> str:
    h, lum, s = _rgb_hex_to_ms_hls(base_rrggbb)
    lum2 = _tint_luminance(tint, lum)
    return _ms_hls_to_rgb_hex(h, lum2, s)


def get_theme_palette_rrggbb(wb) -> list[str] | None:
    """
    Paleta del libro: lt1, dk1, lt2, dk2, accent1..6 como RRGGBB.
    Sin tema cargado (poco habitual) devuelve None.
    """
    lt = getattr(wb, "loaded_theme", None)
    if not lt:
        return None
    try:
        from openpyxl.xml.functions import QName, fromstring

        ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
        root = fromstring(lt)
        te = root.find(QName(ns, "themeElements").text)
        if te is None:
            return None
        schemes = te.findall(QName(ns, "clrScheme").text)
        if not schemes:
            return None
        scheme = schemes[0]
        out: list[str] = []
        for name in (
            "lt1",
            "dk1",
            "lt2",
            "dk2",
            "accent1",
            "accent2",
            "accent3",
            "accent4",
            "accent5",
            "accent6",
        ):
            node = scheme.find(QName(ns, name).text)
            rgb = "000000"
            if node is not None:
                for child in node:
                    tag = child.tag.split("}")[-1]
                    if tag == "sysClr" and "lastClr" in child.attrib:
                        rgb = child.attrib["lastClr"].upper()
                    elif tag == "srgbClr" and "val" in child.attrib:
                        rgb = child.attrib["val"].upper()
            rgb = rgb.replace("#", "")
            out.append(rgb[-6:] if len(rgb) >= 6 else "000000")
        return out
    except (AttributeError, TypeError, ValueError, IndexError):
        return None


def _color_to_rrggbb(color, theme_palette: list[str] | None) -> str | None:
    """Un Color openpyxl → RRGGBB (o None si es automático / sin dato útil)."""
    if color is None:
        return None
    try:
        ct = getattr(color, "type", None) or "rgb"
        if ct == "auto":
            return None
        if ct == "rgb" and color.rgb:
            s = str(color.rgb).upper().replace("0X", "")
            if len(s) == 8:
                return s[2:8]
            if len(s) == 6:
                return s
        if ct == "indexed" and color.indexed is not None:
            i = int(color.indexed)
            if 0 <= i < len(XL_COLORS.COLOR_INDEX):
                raw = XL_COLORS.COLOR_INDEX[i]
                if isinstance(raw, str) and len(raw) >= 8:
                    return raw[2:8].upper()
        if (
            ct == "theme"
            and theme_palette
            and color.theme is not None
        ):
            idx = int(color.theme)
            if 0 <= idx < len(theme_palette):
                tint = float(color.tint) if color.tint is not None else 0.0
                return _theme_base_and_tint_to_rrggbb(theme_palette[idx], tint)
    except (AttributeError, TypeError, ValueError):
        return None
    return None


def _pattern_fill_color_order(f) -> list:
    """Orden de colores a probar: en relleno sólido Excel el color visible suele ir en bgColor."""
    fg = getattr(f, "fgColor", None)
    bg = getattr(f, "bgColor", None)
    ptype = getattr(f, "patternType", None) or getattr(f, "fill_type", None)
    if ptype in (None, "solid"):
        return [bg, fg]
    return [fg, bg]


def iter_cell_fill_rrggbb_candidates(cell, theme_palette: list[str] | None):
    """RRGGBB distintos del relleno (bg antes que fg en sólidos)."""
    f = cell.fill
    if f is None:
        return
    seen: set[str] = set()

    stops = getattr(f, "stop", None)
    if stops:
        for st in stops:
            rgb = _color_to_rrggbb(st.color, theme_palette)
            if rgb and rgb not in seen:
                seen.add(rgb)
                yield rgb
        return

    if hasattr(f, "fgColor"):
        for col in _pattern_fill_color_order(f):
            rgb = _color_to_rrggbb(col, theme_palette)
            if rgb and rgb not in seen:
                seen.add(rgb)
                yield rgb


def phoenix_hue_fallback_checklist(rgb6: str) -> str | None:
    """
    Leyenda Phoenix: CHECKLIST (CHC) = fondo amarillo. Tras tema/tinte el hex rara vez es FFFF00;
    si no hay entrada en cellFillRgbToType, el matiz ~amarillo sigue siendo lista de chequeo.
    (Naranja PROD ~h<0.10; no mezclar.)
    """
    if len(rgb6) != 6:
        return None
    try:
        r = int(rgb6[0:2], 16) / 255.0
        g = int(rgb6[2:4], 16) / 255.0
        b = int(rgb6[4:6], 16) / 255.0
    except ValueError:
        return None
    h, s, v = rgb_to_hsv(r, g, b)
    if v < 0.18 or s < 0.05:
        return None
    # Amarillo / oro checklist (h ~0.17 puro). Naranja PROD (p. ej. #FFA500) h~0.108 → excluido (h < 0.118).
    if 0.118 <= h <= 0.22 and s >= 0.06 and v >= 0.32:
        return "LISTA_CHEQUEO"
    return None


def movement_type_from_cell_fill(
    cell, theme_palette: list[str] | None, fill_rgb_map: dict[str, str]
) -> str | None:
    """
    Primer tipo G-NEEX que coincida con algún canal de color.
    Si hay varios canales y el primero es negro mapeado a ESPECIAL, se prueba el siguiente
    (Excel a veces deja fg en negro y el color real en bg).
    Si no hay hex en el mapa, intenta matiz amarillo (leyenda CHECKLIST).
    """
    cands = list(iter_cell_fill_rrggbb_candidates(cell, theme_palette))
    for rgb6 in cands:
        t = fill_rgb_map.get(rgb6)
        if not t:
            continue
        if t == "ESPECIAL" and rgb6 == "000000" and len(cands) > 1:
            continue
        return t
    for rgb6 in cands:
        t = phoenix_hue_fallback_checklist(rgb6)
        if t:
            return t
    return None


def dump_fill_statistics(
    ws,
    first_move_col: int,
    first_data: int,
    theme_palette: list[str] | None,
    fill_rgb_map: dict[str, str],
    out_path: Path,
) -> None:
    """Lista hex RRGGBB (primer color resuelto por celda) en celdas con ____, para calibrar el mapa."""
    cnt: Counter[str] = Counter()
    for c in range(first_move_col, ws.max_column + 1):
        for r in range(first_data, ws.max_row + 1):
            cell_obj = ws.cell(r, c)
            v = cell_obj.value
            if not isinstance(v, str) or "____" not in v:
                continue
            primary = None
            for rgb in iter_cell_fill_rrggbb_candidates(cell_obj, theme_palette):
                primary = rgb
                break
            cnt[primary or "(sin color)"] += 1
    lines = [
        "# Primer color resuelto por celda (bg antes que fg en relleno sólido).",
        "# Copia los hex \"(sin mapear)\" a cellFillRgbToType en phoenix-type-map.json",
        "",
    ]
    for hexv, n in cnt.most_common():
        if hexv == "(sin color)":
            lines.append(f"{hexv}\t{n}")
            continue
        mapped = fill_rgb_map.get(hexv)
        tag = f"\t-> {mapped}" if mapped else "\t-> (sin mapear)"
        lines.append(f"{hexv}\t{n}{tag}")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def match_header_to_type(hdr: str, type_map: dict[str, str]) -> str | None:
    """
    Coincide etiquetas de subcabecera aunque vengan con espacios/guiones extra
    (p. ej. RECEPTION CHANTIER antes que RECEPTION por longitud de clave).
    """
    flat_hdr = re.sub(r"[^A-Z0-9]", "", hdr.upper())
    if len(flat_hdr) < 4:
        return None
    for k in sorted(type_map.keys(), key=len, reverse=True):
        if len(k) < 4:
            continue
        fk = re.sub(r"[^A-Z0-9]", "", k.upper())
        if len(fk) >= 4 and fk in flat_hdr:
            return type_map[k]
    return None


def normalize_hdr_label(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return None
    s = str(val).strip().upper().replace(" ", "_")
    return s if s else None


def is_project_only_header(val) -> bool:
    """Subcabeceras que son solo nº de proyecto: no son tipo ni ESPECIAL."""
    if val is None:
        return False
    if isinstance(val, bool):
        return False
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        s = val.strip()
        return bool(s) and s.isdigit()
    return False


def column_movement_types(
    ws, header_row: int, first_move_col: int, type_map: dict[str, str], default_type: str
) -> dict[int, str]:
    """
    En Phoenix **cada columna** (desde first_move_col) corresponde a un movimiento.
    La fila de subcabecera mezcla etiquetas (ADJUST, RECEPTION, ASSEMBLAGE…) y **números
    de proyecto**: el dato con `____` suele estar bajo el número, y la etiqueta de tipo
    va **a la izquierda** (p. ej. columna 10 = ADJUST, datos en 12–13 con cabecera numérica).

    Por eso el tipo se obtiene **mirando hacia la izquierda** desde cada columna hasta
    encontrar la última etiqueta reconocida en `type_map`, o texto libre → ESPECIAL.
    Si se encuentra INITIAL_QUANTITY en ese barrido, la columna se trata como AJUSTE
    (snapshot / columna inicial). Si no hay nada útil, `default_type`.
    """
    max_c = ws.max_column + 1
    col_types: dict[int, str] = {}

    def type_for_col(c: int) -> str:
        for cc in range(c, first_move_col - 1, -1):
            hdr = effective_cell_value(ws, header_row, cc)
            nk = normalize_hdr_label(hdr) if isinstance(hdr, str) else None

            if nk == "INITIAL_QUANTITY":
                return "AJUSTE"

            if is_project_only_header(hdr):
                continue

            if nk and nk in type_map:
                return type_map[nk]

            if isinstance(hdr, str) and str(hdr).strip():
                mt = match_header_to_type(hdr, type_map)
                if mt:
                    return mt
                return "ESPECIAL"

        return default_type

    for c in range(first_move_col, max_c):
        col_types[c] = type_for_col(c)
    return col_types


def discover_date_row(ws, header_row: int) -> int:
    if header_row > 1:
        return header_row - 1
    return 8


def main() -> None:
    ap = argparse.ArgumentParser(description="HISTORY xlsx → CSV inventario + JSON respaldo (separados)")
    ap.add_argument("--input", "-i", required=True, help="Ruta Libro1.xlsx")
    ap.add_argument(
        "--gneex-backup-json",
        default="",
        help="Respaldo JSON G-NEEX: alinea inventario completo (stocks, ubicación, id, cajas…) con ese snapshot final.",
    )
    ap.add_argument(
        "--gneex-inventory-csv",
        default="",
        help="Tras --gneex-backup-json: completa ubicación, lotes, min/max, etc. desde el CSV exportado "
        "(no cambia cantidades). Sin respaldo: solo metadatos desde CSV.",
    )
    ap.add_argument("--sheet", default="HISTORY")
    ap.add_argument(
        "--out-dir",
        default="",
        help="Carpeta salida (default: migration/generated)",
    )
    ap.add_argument("--no-node", action="store_true", help="No ejecutar build-gneex-backup.mjs")
    ap.add_argument("--header-row", type=int, default=0)
    ap.add_argument("--date-row", type=int, default=0, help="Fila Excel de fechas (0 = header_row-1)")
    ap.add_argument("--code-col", type=int, default=3)
    ap.add_argument("--desc-col", type=int, default=4)
    ap.add_argument("--cat-col", type=int, default=5)
    ap.add_argument(
        "--loc-col",
        type=int,
        default=6,
        help="Columna Excel de ubicación (0 = omitir). Por defecto 6 (entre categoría e INITIAL QUANTITY en 7).",
    )
    ap.add_argument("--initial-col", type=int, default=7)
    ap.add_argument(
        "--last-initial-col",
        type=int,
        default=0,
        help="Columna INITIAL QUANTITY inventario (0 = la más a la izquierda / más reciente)",
    )
    ap.add_argument("--first-move-col", type=int, default=9)
    ap.add_argument(
        "--dump-fill-stats",
        action="store_true",
        help="Escribe phoenix_fill_color_stats.txt con hex de relleno en celdas ____ y termina (calibrar cellFillRgbToType)",
    )
    ap.add_argument(
        "--audit-reconcile",
        action="store_true",
        help="Auditoría: mismatches si suma(deltas)+snapshots ≠ inventario (INITIAL QUANTITY más a la izquierda)",
    )
    ap.add_argument(
        "--audit-tolerance",
        type=float,
        default=0.5,
        help="Tolerancia |calculado - inventario| en --audit-reconcile (default 0.5)",
    )
    args = ap.parse_args()

    migration_dir = Path(__file__).resolve().parent
    out_dir = Path(args.out_dir).resolve() if args.out_dir else (migration_dir / "generated")
    out_dir.mkdir(parents=True, exist_ok=True)

    in_path = Path(args.input).expanduser().resolve()
    if not in_path.is_file():
        print(f"No existe: {in_path}", file=sys.stderr)
        sys.exit(1)

    type_map, default_type, fill_rgb_map = load_type_map(migration_dir)

    # data_only=True conserva estilos (fill); necesario para color de celdas con ____
    wb = openpyxl.load_workbook(in_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        print("Hojas:", wb.sheetnames, file=sys.stderr)
        sys.exit(1)
    ws = wb[args.sheet]
    theme_palette = get_theme_palette_rrggbb(wb)

    header_row = args.header_row if args.header_row > 0 else find_header_row(ws)
    date_row = args.date_row if args.date_row > 0 else discover_date_row(ws, header_row)
    first_data = header_row + 1
    if args.last_initial_col > 0:
        last_initial_col = args.last_initial_col
    else:
        last_initial_col = find_last_initial_quantity_column(ws, header_row)
        if last_initial_col is None:
            print(
                "No se encontró subcabecera INITIAL_QUANTITY. "
                "Indica la columna con --last-initial-col N (número de columna Excel).",
                file=sys.stderr,
            )
            sys.exit(1)
    print(
        f"Inventario: columna INITIAL_QUANTITY (más a la izquierda / más reciente) = {last_initial_col}"
    )
    col_dt = column_dates(ws, date_row)
    col_mov_type = column_movement_types(ws, header_row, args.first_move_col, type_map, default_type)

    if args.audit_reconcile:
        (
            mism,
            col_hits,
            total_art,
            ok_n,
            mal_leidas,
            ops_descuadre_todas,
        ) = run_reconciliation_audit(
            ws,
            header_row,
            first_data,
            last_initial_col,
            args.first_move_col,
            args.initial_col,
            args.code_col,
            col_dt,
            abs_tol=args.audit_tolerance,
        )
        audit_csv = out_dir / "phoenix_reconcile_mismatches.csv"
        audit_txt = out_dir / "phoenix_reconcile_by_column.txt"
        audit_mal = out_dir / "phoenix_reconcile_mal_leidas.csv"
        audit_ops = out_dir / "phoenix_reconcile_ops_descuadre.csv"
        with audit_csv.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(
                f,
                fieldnames=[
                    "fila",
                    "codigo",
                    "initial_col",
                    "calculado",
                    "inventario",
                    "diff",
                    "movimientos_col_delta",
                    "detalle_modos_parseo",
                    "alertas_lectura",
                ],
            )
            w.writeheader()
            w.writerows(mism)
        with audit_mal.open("w", encoding="utf-8-sig", newline="") as f:
            wm = csv.DictWriter(
                f,
                fieldnames=[
                    "fila",
                    "codigo",
                    "col_movimiento",
                    "delta",
                    "modo_parseo",
                    "trozo_celda",
                    "motivo",
                ],
            )
            wm.writeheader()
            wm.writerows(mal_leidas)
        with audit_ops.open("w", encoding="utf-8-sig", newline="") as f:
            wo = csv.DictWriter(
                f,
                fieldnames=[
                    "fila",
                    "codigo",
                    "col_movimiento",
                    "delta",
                    "modo_parseo",
                    "trozo_celda",
                    "diff_inventario_fila",
                ],
            )
            wo.writeheader()
            wo.writerows(ops_descuadre_todas)
        diag_path = out_dir / "phoenix_reconcile_diagnostico_filas.txt"
        diag_path.write_text(
            diagnostico_texto_filas_problematicas(mism, mal_leidas, ops_descuadre_todas),
            encoding="utf-8",
        )
        lines = [
            "# Ver también: phoenix_reconcile_mal_leidas.csv (legacy / sin parse) y",
            "# phoenix_reconcile_ops_descuadre.csv (todas las ops leídas en filas descuadradas).",
            "# Columnas de MOVIMIENTO con ____ en filas de artículo descuadradas:",
            f"# Filas revisadas: {total_art} | OK: {ok_n} | descuadradas: {len(mism)} | sospechosas: {len(mal_leidas)}",
            "",
        ]
        for c, n in col_hits.most_common():
            lines.append(
                f"col {c}\t{n} filas de artículo descuadradas con delta en esta columna de movimiento"
            )
        audit_txt.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(f"Auditoría concordancia inventario: {audit_csv}")
        print(f"Operaciones sospechosas / mal leídas: {audit_mal}")
        print(f"Todas las operaciones en filas descuadradas: {audit_ops}")
        print(f"Diagnóstico legible por fila problemática: {diag_path}")
        print(f"Resumen por columna: {audit_txt}")

    if args.dump_fill_stats:
        stats_path = out_dir / "phoenix_fill_color_stats.txt"
        dump_fill_statistics(
            ws, args.first_move_col, first_data, theme_palette, fill_rgb_map, stats_path
        )
        print(f"Estadísticas de relleno (celdas con ____): {stats_path}")
        sys.exit(0)

    # —— Inventario (último conteo; última fila gana por código) ——
    inv_order: list[str] = []
    inv_by_code: dict[str, dict] = {}

    # —— Inventario: todas las filas; última fila por código gana ——
    for r in range(first_data, ws.max_row + 1):
        raw_code = effective_cell_value(ws, r, args.code_col)
        if raw_code is None or str(raw_code).strip() == "":
            continue
        code = str(raw_code).strip()
        if code.upper() == "CODE":
            continue

        desc = effective_cell_value(ws, r, args.desc_col)
        cat = effective_cell_value(ws, r, args.cat_col)
        desc_s = str(desc).strip() if desc is not None else ""
        cat_s = str(cat).strip() if cat is not None else ""

        loc_s = ""
        if getattr(args, "loc_col", 0) and args.loc_col > 0:
            raw_loc = effective_cell_value(ws, r, args.loc_col)
            if raw_loc is not None and str(raw_loc).strip() != "":
                loc_s = str(raw_loc).strip()

        try:
            ini = float(effective_cell_value(ws, r, args.initial_col) or 0)
        except (TypeError, ValueError):
            ini = 0.0

        main_s = stock_from_initial_quantity_cell(
            effective_cell_value(ws, r, last_initial_col), ini
        )
        key = code.lower()

        notes = (
            "Origen: Phoenix HISTORY. StockPrincipal = celda bajo INITIAL_QUANTITY "
            f"más a la izquierda (más reciente; col {last_initial_col}); si vacío, --initial-col."
        )
        if main_s < 0:
            notes += " ⚠ Stock negativo; revisar."
        if key in inv_by_code:
            notes += " Código duplicado: prevalece esta fila."

        inv_by_code[key] = {
            "code": code,
            "description": desc_s,
            "category": cat_s,
            "mainStock": main_s,
            "prodStock": 0,
            "transStock": 0,
            "location": loc_s,
            "notes": notes,
        }
        if key not in inv_order:
            inv_order.append(key)

    inv_list = [inv_by_code[k] for k in inv_order]

    merge_backup_stats: dict[str, int] | None = None
    merge_backup_path: str | None = None
    if getattr(args, "gneex_backup_json", "").strip():
        bk = Path(args.gneex_backup_json.strip()).expanduser().resolve()
        if not bk.is_file():
            print(f"No existe --gneex-backup-json: {bk}", file=sys.stderr)
            sys.exit(1)
        inv_list, merge_backup_stats = merge_inv_list_stocks_from_gneex_backup(inv_list, bk)
        merge_backup_path = bk.name
        print(
            f"Respaldo G-NEEX final (stocks + ubicación, etc.): {merge_backup_stats['from_backup_final']} códigos; "
            f"solo Phoenix (sin fila en respaldo): {merge_backup_stats['phoenix_only']}"
        )

    merge_csv_stats: dict[str, int] | None = None
    merge_csv_path: str | None = None
    if getattr(args, "gneex_inventory_csv", "").strip():
        merge_p = Path(args.gneex_inventory_csv.strip()).expanduser().resolve()
        if not merge_p.is_file():
            print(f"No existe --gneex-inventory-csv: {merge_p}", file=sys.stderr)
            sys.exit(1)
        inv_list, merge_csv_stats = merge_inv_list_from_gneex_csv(inv_list, merge_p)
        merge_csv_path = merge_p.name
        if getattr(args, "gneex_backup_json", "").strip():
            print(
                f"Complemento CSV (ubicación, lotes, min/max…; stocks siguen del respaldo): "
                f"{merge_csv_stats['metadata_from_csv']} códigos; sin fila CSV: {merge_csv_stats['no_csv_row']}"
            )
        else:
            print(
                f"Metadatos desde CSV G-NEEX: {merge_csv_stats['metadata_from_csv']} códigos; "
                f"sin fila CSV: {merge_csv_stats['no_csv_row']}"
            )

    # —— Movimientos: una columna = un movimiento; filas = líneas de artículo ——
    raw_movements: list[dict] = []

    for c in range(args.first_move_col, ws.max_column + 1):
        lines_by_code: dict[str, dict] = {}
        project_ids: list[str] = []
        dates_seen: list[str] = []
        sample_note: list[str] = []
        line_fill_types: list[str] = []

        for r in range(first_data, ws.max_row + 1):
            raw_code = effective_cell_value(ws, r, args.code_col)
            if raw_code is None or str(raw_code).strip() == "":
                continue
            code = str(raw_code).strip()
            if code.upper() == "CODE":
                continue

            cell_obj = anchor_cell(ws, r, c)
            cell = cell_obj.value
            if not isinstance(cell, str) or "____" not in cell:
                continue
            parsed = parse_movement_cell(cell, col_dt.get(c))
            if not parsed:
                continue
            ft = movement_type_from_cell_fill(cell_obj, theme_palette, fill_rgb_map)
            if ft:
                line_fill_types.append(ft)
            qty, date_s, project_id, raw = parsed
            ck = code.lower()
            lines_by_code[ck] = {
                "code": code,
                "quantity": float(qty),
                "target": "main",
            }
            if project_id:
                project_ids.append(project_id)
            dates_seen.append(date_s)
            if len(sample_note) < 4:
                sample_note.append(f"{code}:{qty}")

        if not lines_by_code:
            continue

        lines = list(lines_by_code.values())
        if line_fill_types:
            mov_type = Counter(line_fill_types).most_common(1)[0][0]
        else:
            mov_type = col_mov_type.get(c, default_type)
        col_date = col_dt.get(c)

        if dates_seen:
            date_s = min(dates_seen)
        elif col_date:
            date_s = col_date
        else:
            date_s = "1970-01-01"

        if project_ids:
            proj = max(set(project_ids), key=project_ids.count)
        else:
            proj = ""

        notes = (
            f"Phoenix Excel col {c} | {len(lines)} artículos | "
            + "; ".join(sample_note)
            + ("…" if len(lines) > len(sample_note) else "")
        )

        raw_movements.append(
            {
                "type": mov_type,
                "projectId": proj,
                "date": f"{date_s}T12:00:00.000Z",
                "notes": notes[:500],
                "lines": lines,
                "_sort": (date_s, c),
            }
        )

    raw_movements.sort(key=lambda x: x["_sort"])

    # intermediate.json para build-gneex-backup.mjs
    movements_out: list[dict] = []
    for i, m in enumerate(raw_movements, start=1):
        movements_out.append(
            {
                "reference": str(i).zfill(8),
                "date": m["date"],
                "type": m["type"],
                "projectId": m["projectId"],
                "notes": m["notes"],
                "createdBy": "Migración Phoenix (HISTORY xlsx)",
                "lines": m["lines"],
            }
        )

    intermediate = {
        "meta": {
            "source": "Phoenix HISTORY xlsx",
            "inputFile": in_path.name,
            "sheet": args.sheet,
            "separateInventoryCsv": "GNEEX_Inventario_PHOENIX.csv",
            "lastInitialQuantityColumn": last_initial_col,
            "mainStockRule": "mainStock = cell under leftmost INITIAL QUANTITY subheader (most recent; lastInitialQuantityColumn); empty → --initial-col.",
            "gneexBackupJsonMerge": merge_backup_path,
            "gneexBackupJsonMergeStats": merge_backup_stats,
            "gneexInventoryCsvMerge": merge_csv_path,
            "gneexInventoryCsvMergeStats": merge_csv_stats,
            "movementCount": len(movements_out),
            "movementLineCount": sum(len(m.get("lines") or []) for m in movements_out),
            "itemCount": len(inv_list),
            "phoenixModel": "one column = one movement; rows = item lines",
        },
        "inventory": inv_list,
        "movements": movements_out,
    }

    intermediate_path = out_dir / "intermediate_from_HISTORY.json"
    intermediate_path.write_text(json.dumps(intermediate, ensure_ascii=False, indent=2), encoding="utf-8")

    # CSV inventario (Ids alineados con orden del intermediate)
    csv_path = out_dir / "GNEEX_Inventario_PHOENIX.csv"
    csv_rows: list[dict] = []
    for idx, it in enumerate(inv_list, start=1):
        pid = it.get("id")
        row_id = str(pid).strip() if pid not in (None, "") else str(idx)
        csv_rows.append(
            {
                "Codigo": it["code"],
                "Descripcion": it["description"],
                "Categoria": it["category"],
                "StockPrincipal": it["mainStock"],
                "StockProduccion": it.get("prodStock", 0) or 0,
                "StockTransformacion": it.get("transStock", 0) or 0,
                "CantidadPorCaja": it.get("qtyPerBox", 0) or 0,
                "NumeroCajas": it.get("numBoxes", 0) or 0,
                "Ubicacion": it.get("location") or "",
                "FechaExpedicion": it.get("expDate") or "",
                "DiasParaExpirar": ""
                if it.get("daysToExpire") is None
                else str(it.get("daysToExpire")),
                "FechaExpiracion": it.get("expirationDate") or "",
                "Proveedor": it.get("supplier") or "",
                "UltimaOrden": it.get("lastOrder") or "",
                "Detalles": it.get("details") or "",
                "Id": row_id,
                "StockMinimo": it.get("minStock", 0) or 0,
                "StockMaximo": it.get("maxStock", 0) or 0,
                "VidaUtilMeses": it.get("shelfLifeMonths", 0) or 0,
                "Notas": it.get("notes", ""),
                "LotesJson": json.dumps(it["expirations"], ensure_ascii=False)
                if isinstance(it.get("expirations"), list)
                else "[]",
            }
        )
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_HEADERS, extrasaction="ignore")
        w.writeheader()
        w.writerows(csv_rows)

    backup_out = out_dir / "GNEEX_Respaldo_PHOENIX.json"
    if args.no_node:
        print("— intermediate:", intermediate_path)
        print("— inventario CSV:", csv_path)
        total_lines = sum(len(m.get("lines") or []) for m in movements_out)
        print(f"Artículos: {len(csv_rows)} | Movimientos (columnas): {len(movements_out)} | Líneas: {total_lines}")
        print("Omitido Node. Ejecuta:")
        print(
            f'  node "{migration_dir / "build-gneex-backup.mjs"}" '
            f'--input "{intermediate_path}" --output "{backup_out}"'
        )
        return

    node_cmd = ["node", str(migration_dir / "build-gneex-backup.mjs"), "--input", str(intermediate_path), "--output", str(backup_out)]
    pr = subprocess.run(node_cmd, cwd=str(migration_dir), shell=False)
    if pr.returncode != 0:
        print("Error ejecutando Node (¿tienes Node instalado?).", file=sys.stderr)
        sys.exit(pr.returncode or 1)
    print("— inventario CSV:", csv_path)
    print("— respaldo JSON:", backup_out)
    print("— intermediate (auditoría):", intermediate_path)
    total_lines = sum(len(m.get("lines") or []) for m in movements_out)
    print(f"Artículos: {len(csv_rows)} | Movimientos (columnas): {len(movements_out)} | Líneas: {total_lines}")


if __name__ == "__main__":
    main()
