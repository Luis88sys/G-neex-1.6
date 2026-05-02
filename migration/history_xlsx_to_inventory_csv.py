#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inventario importable (CSV G-NEEX) a partir de la hoja Phoenix HISTORY (Libro1.xlsx).

Regla: **StockPrincipal** = valor bajo la columna **INITIAL QUANTITY** de inventario
(**más a la izquierda** / más reciente; misma lógica que el respaldo JSON). Si la celda está vacía,
se usa el valor de **--initial-col** como respaldo.

Uso:
  py -3 migration/history_xlsx_to_inventory_csv.py --input "%USERPROFILE%\\Desktop\\Libro1.xlsx"
  py -3 migration/history_xlsx_to_inventory_csv.py --input libro.xlsx --output migration/generated/inventario.csv

Requiere: pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl: py -3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from gneex_inventory_csv_merge import (
    merge_flat_csv_rows_from_gneex_csv,
    merge_flat_csv_rows_stocks_from_gneex_backup,
)
from phoenix_sheet_utils import (
    effective_cell_value,
    find_last_initial_quantity_column,
    stock_from_initial_quantity_cell,
)

HEADERS = [
    "Codigo",
    "Descripcion",
    "Categoria",
    "StockPrincipal",
    "StockProduccion",
    "StockTransformacion",
    "CantidadPorCaja",
    "NumeroCajas",
    "Ubicacion",
    "FechaExpedicion",
    "DiasParaExpirar",
    "FechaExpiracion",
    "Proveedor",
    "UltimaOrden",
    "Detalles",
    "Id",
    "StockMinimo",
    "StockMaximo",
    "VidaUtilMeses",
    "Notas",
    "LotesJson",
]

def find_header_row(ws, max_scan: int = 30) -> int:
    for r in range(1, max_scan + 1):
        v = ws.cell(r, 3).value
        if v is None:
            continue
        if str(v).strip().upper() == "CODE":
            return r
    raise ValueError("No se encontró fila con 'CODE' en columna C (ajusta --header-row).")


def main() -> None:
    ap = argparse.ArgumentParser(description="Phoenix HISTORY → CSV inventario G-NEEX (último conteo)")
    ap.add_argument("--input", "-i", required=True, help="Ruta al .xlsx (p. ej. Libro1.xlsx)")
    ap.add_argument("--sheet", default="HISTORY", help="Nombre de la hoja (default: HISTORY)")
    ap.add_argument(
        "--output",
        "-o",
        default="",
        help="Ruta del CSV (default: migration/generated/GNEEX_Inventory_Phoenix_HISTORY.csv)",
    )
    ap.add_argument("--header-row", type=int, default=0, help="Fila del encabezado CODE (0 = autodetectar)")
    ap.add_argument("--code-col", type=int, default=3)
    ap.add_argument("--desc-col", type=int, default=4)
    ap.add_argument("--cat-col", type=int, default=5)
    ap.add_argument(
        "--loc-col",
        type=int,
        default=6,
        help="Columna Excel de ubicación (0 = omitir). Por defecto 6.",
    )
    ap.add_argument("--initial-col", type=int, default=7)
    ap.add_argument(
        "--last-initial-col",
        type=int,
        default=0,
        help="Columna INITIAL QUANTITY inventario (0 = la más a la izquierda / más reciente)",
    )
    ap.add_argument(
        "--gneex-backup-json",
        default="",
        help="Respaldo JSON G-NEEX: cantidades desde phoenix-inventory (no el CSV).",
    )
    ap.add_argument(
        "--gneex-inventory-csv",
        default="",
        help="CSV G-NEEX: solo metadatos (ubicación, id, cajas…); no cantidades.",
    )
    args = ap.parse_args()

    in_path = Path(args.input).expanduser().resolve()
    if not in_path.is_file():
        print(f"No existe el archivo: {in_path}", file=sys.stderr)
        sys.exit(1)

    out_default = Path(__file__).resolve().parent / "generated" / "GNEEX_Inventory_Phoenix_HISTORY.csv"
    out_path = Path(args.output).expanduser().resolve() if args.output else out_default
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(in_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"Hojas disponibles: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb[args.sheet]

    header_row = args.header_row if args.header_row > 0 else find_header_row(ws)
    first_data_row = header_row + 1
    if args.last_initial_col > 0:
        last_initial_col = args.last_initial_col
    else:
        last_initial_col = find_last_initial_quantity_column(ws, header_row)
        if last_initial_col is None:
            print(
                "No se encontró subcabecera INITIAL_QUANTITY. "
                "Usa --last-initial-col N (número de columna Excel).",
                file=sys.stderr,
            )
            sys.exit(1)
    print(
        f"Inventario: columna INITIAL_QUANTITY (más a la izquierda / más reciente) = {last_initial_col}"
    )

    by_code: dict[str, dict] = {}
    order: list[str] = []

    for r in range(first_data_row, ws.max_row + 1):
        raw_code = effective_cell_value(ws, r, args.code_col)
        if raw_code is None or str(raw_code).strip() == "":
            continue
        code = str(raw_code).strip()
        if code.upper() == "CODE":
            continue

        desc = effective_cell_value(ws, r, args.desc_col)
        cat = effective_cell_value(ws, r, args.cat_col)
        desc_s = str(desc).strip() if desc is not None else ""
        cat_s = str(cat).strip() if cat is not None else ""

        loc_s = ""
        if getattr(args, "loc_col", 0) and args.loc_col > 0:
            raw_loc = effective_cell_value(ws, r, args.loc_col)
            if raw_loc is not None and str(raw_loc).strip() != "":
                loc_s = str(raw_loc).strip()

        init_cell = effective_cell_value(ws, r, args.initial_col)
        try:
            initial = float(init_cell) if init_cell is not None else 0.0
        except (TypeError, ValueError):
            initial = 0.0

        main_stock = stock_from_initial_quantity_cell(
            effective_cell_value(ws, r, last_initial_col), initial
        )
        key = code.lower()

        notes_parts = [
            "Origen: Phoenix HISTORY. StockPrincipal = INITIAL_QUANTITY más a la izquierda (más reciente) "
            f"(col {last_initial_col}); si vacío, --initial-col."
        ]
        if main_stock < 0:
            notes_parts.append("⚠ Stock negativo; revisar fila en Excel.")
        if key in by_code:
            notes_parts.append("Código repetido: prevalece esta fila (la última en el archivo).")

        row_obj = {
            "Codigo": code,
            "Descripcion": desc_s,
            "Categoria": cat_s,
            "StockPrincipal": main_stock,
            "StockProduccion": 0,
            "StockTransformacion": 0,
            "CantidadPorCaja": 0,
            "NumeroCajas": 0,
            "Ubicacion": loc_s,
            "FechaExpedicion": "",
            "DiasParaExpirar": "",
            "FechaExpiracion": "",
            "Proveedor": "",
            "UltimaOrden": "",
            "Detalles": "",
            "Id": "",
            "StockMinimo": 0,
            "StockMaximo": 0,
            "VidaUtilMeses": 0,
            "Notas": " ".join(notes_parts),
            "LotesJson": "[]",
        }

        if key not in by_code:
            order.append(key)
        by_code[key] = row_obj

    rows_out = [by_code[k] for k in order]
    for i, row in enumerate(rows_out, start=1):
        if not (row.get("Id") or "").strip():
            row["Id"] = str(i)

    if getattr(args, "gneex_backup_json", "").strip():
        bk = Path(args.gneex_backup_json.strip()).expanduser().resolve()
        if not bk.is_file():
            print(f"No existe --gneex-backup-json: {bk}", file=sys.stderr)
            sys.exit(1)
        rows_out, bk_stats = merge_flat_csv_rows_stocks_from_gneex_backup(rows_out, bk)
        print(
            f"Respaldo G-NEEX final: {bk_stats['from_backup_final']} códigos; "
            f"solo Phoenix: {bk_stats['phoenix_only']}"
        )

    if getattr(args, "gneex_inventory_csv", "").strip():
        merge_p = Path(args.gneex_inventory_csv.strip()).expanduser().resolve()
        if not merge_p.is_file():
            print(f"No existe --gneex-inventory-csv: {merge_p}", file=sys.stderr)
            sys.exit(1)
        rows_out, merge_stats = merge_flat_csv_rows_from_gneex_csv(rows_out, merge_p)
        if getattr(args, "gneex_backup_json", "").strip():
            print(
                f"Complemento CSV (ubicación, etc.; stocks del respaldo): "
                f"{merge_stats['metadata_from_csv']} códigos; sin fila CSV: {merge_stats['no_csv_row']}"
            )
        else:
            print(
                f"Metadatos desde CSV: {merge_stats['metadata_from_csv']} códigos; "
                f"sin fila CSV: {merge_stats['no_csv_row']}"
            )

    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows_out)

    print(f"Escrito: {out_path}")
    print(
        f"Artículos: {len(rows_out)} (stock INITIAL_QUANTITY inventario col {last_initial_col}, "
        "más a la izquierda / más reciente; duplicados: gana la última fila)"
    )


if __name__ == "__main__":
    main()
